"""Excel (.xlsx) parser。

策略：
- 用 openpyxl 读所有 sheet
- 每个 sheet → 一份 ParsedTable（headers 从首行推断）
- sheet 名 → "章节"（TocEntry，level=1）
- 单元格合并、公式等暂忽略（只取值）

Excel 在芯片 spec 中常用于：
- pin table
- balls / package mapping
- ordering info
TableEntityExtractor 可以从中扫描表格。
"""

from __future__ import annotations

from pathlib import Path

from docgraph.graph.schema import (
    ParsedDoc,
    ParsedPage,
    ParsedTable,
    TextBlock,
    TocEntry,
)
from docgraph.parsers.base import ParseContext
from docgraph.parsers.normalize import populate_l0_blocks


class XlsxParser:
    name = "xlsx"
    supports = {".xlsx", ".xlsm"}
    version = "0.1"

    MAX_ROWS_PER_SHEET = 5000

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.supports

    def parse(self, path: Path, ctx: ParseContext) -> ParsedDoc:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:  # pragma: no cover
            raise RuntimeError(
                "openpyxl is a required DocGraph dependency; "
                "reinstall or repair the docgraph-core installation"
            ) from e

        wb = load_workbook(filename=str(path), data_only=True, read_only=True)

        text_blocks: list[TextBlock] = []
        tables: list[ParsedTable] = []
        toc: list[TocEntry] = []
        order = 0

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            # 章节
            toc.append(
                TocEntry(
                    level=1,
                    title=sheet_name,
                    section_path=str(sheet_idx),
                )
            )
            text_blocks.append(
                TextBlock(
                    text=f"Sheet: {sheet_name}",
                    reading_order=order,
                    is_heading=True,
                    heading_level=1,
                )
            )
            order += 1

            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True, max_row=self.MAX_ROWS_PER_SHEET):
                cells = [self._fmt_cell(c) for c in row]
                if not any(cells):
                    continue
                rows.append(cells)
            if not rows:
                continue
            headers = rows[0]
            data_rows = rows[1:] if len(rows) > 1 else []
            tables.append(
                ParsedTable(
                    headers=headers,
                    rows=data_rows,
                    caption=sheet_name,
                )
            )
            # 同时把表格内容拼成 text，便于关键字检索
            joined = "\n".join("\t".join(r) for r in rows[:200])
            text_blocks.append(TextBlock(text=joined, reading_order=order))
            order += 1

        wb.close()

        single_page = ParsedPage(
            page_no=1,
            text_blocks=text_blocks,
            tables=tables,
        )
        populate_l0_blocks(single_page, doc_id=ctx.doc_id, parser=self.name)

        return ParsedDoc(
            doc_id=ctx.doc_id,
            source_path=str(path),
            pages=[single_page],
            metadata=ctx.metadata,
            toc=toc,
            parser=self.name,
            parser_version=self.version,
        )

    @staticmethod
    def _fmt_cell(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            # 去掉无意义的尾零
            if v.is_integer():
                return str(int(v))
        return str(v).strip()
